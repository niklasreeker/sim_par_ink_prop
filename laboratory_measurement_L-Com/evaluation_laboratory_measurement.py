#!/usr/bin/env python3
"""
Evaluation of laboratory measurements for L-Com 5500 / Pico 3000
Master's Thesis Reeker

Directory Structure:

    laboratory_measurement_L-Com/
    +-- evaluation_laboratory_measurement.py  <- this script
    +-- measurement_data/                     <- PLC CSV files
    +-- results/                              <- Excel file and plots

The script:
  1. Reads all CSV files from "measurement_data", calculates the
     composition from initial weights, and exports everything as an
     Excel file to "results".
  2. Generates a plot for each sample ID showing density, sound velocity,
     and temperature over time, including variation bands and markers for
     every formulation/recipe change.

Usage: python evaluation_laboratory_measurement.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

# =====================================================================
# CONFIGURATION
# =====================================================================

BASE_DIR         = Path(__file__).resolve().parent
MEASUREMENT_DATA = BASE_DIR / "measurement_data"
RESULTS_DIR      = BASE_DIR / "results"

# Mass fractions of concentrate SL120 (manufacturer specification).
# If these values change, the entire campaign is recalculated.
SL120 = {"Al": 0.20, "IPA": 0.40, "PG": 0.40}

# Weighing columns used to determine the composition
INITIAL_WEIGHTS = ["m_SL120", "m_Wasser", "m_IPA", "m_PG", "m_MG"]

# Threshold for mass change [g] to consider a recipe/formulation modified.
# Setting this below balance readability would be meaningless.
CHANGE_THRESHOLD = 0.05

# Fluctuation range: factor for the central 90% confidence/coverage interval.
# For approximately normally distributed single measurements, 90% of values
# lie within mean +/- 1.645 * standard deviation (5th to 95th percentile).
Z_90 = 1.645

# --- Filter for unusable records ------------------------------------
# A record written while the statistics buffer was still empty (N = 0)
# contains nothing but zeros. This happens when the automatic write
# fires shortly after ResetStatistik. Such records are excluded from
# the plots but kept in the Excel export, flagged with a reason.
MIN_SAMPLES = 1

# Physically plausible ranges. A value outside these limits means the
# sensor was not measuring a liquid (air, station failure, NaN).
VALUE_RANGES = {
    "Rho_M": (500.0, 2000.0),    # kg/m3
    "C_M":   (1000.0, 2200.0),   # m/s
    "T_M":   (-10.0, 100.0),     # degC
}

COLORS = {
    "rho": "#1f4e79",
    "c":   "#a33b3b",
    "T":   "#3f7d5a",
    "Al":  "#8c4a2f",
    "IPA": "#3d6b8c",
    "PG":  "#5c8c5c",
    "MG":  "#8c7a3d",
    "grid": "#cccccc",
    "mark": "#666666",
}

COLUMN_LABELS = {
    "m_SL120": "SL120", "m_Wasser": "Water",
    "m_IPA": "IPA", "m_PG": "PG", "m_MG": "MG",
}


# =====================================================================
# 1  DATA IMPORT AND PREPROCESSING
# =====================================================================

def load_measurement_data() -> pd.DataFrame:
    """Reads all CSV files from the measurement data directory."""
    files = sorted(MEASUREMENT_DATA.glob("*.csv"))
    if not files:
        sys.exit(f"No CSV files found in {MEASUREMENT_DATA}.")

    parts = []
    for f in files:
        # comment="/" removes trailing comment lines such as //END,
        # skipinitialspace strips leading whitespace from values
        d = pd.read_csv(f, comment="/", skipinitialspace=True)
        d.columns = [c.strip() for c in d.columns]
        d["Source_File"] = f.name
        parts.append(d)
        print(f"  {f.name}: {len(d)} records")

    return pd.concat(parts, ignore_index=True)


def add_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Constructs a timestamp and calculates elapsed time per sample.

    If the PLC/CPU clock is incorrect, absolute timestamps are inaccurate,
    but relative time differences remain valid. Therefore, elapsed time since
    the start of each sample is used for all plots.
    """
    df = df.copy()
    df["Timestamp"] = pd.NaT

    if {"Date", "UTC Time"}.issubset(df.columns):
        # S7 outputs ISO date and time with milliseconds,
        # e.g., "2026-08-25" and "09:00:00.000".
        raw = (df["Date"].astype(str).str.strip() + " "
               + df["UTC Time"].astype(str).str.strip().str.upper())
        ts = pd.to_datetime(raw, format="%Y-%m-%d %H:%M:%S.%f",
                            errors="coerce")
        if ts.isna().any():   # fallback for alternate CPU date/time formats
            ts = pd.to_datetime(raw, errors="coerce")
        df["Timestamp"] = ts

    if df["Timestamp"].notna().any():
        years = df["Timestamp"].dt.year.dropna()
        if not years.empty and years.min() < 2024:
            print(f"  Note: CPU clock is set to {int(years.min())}. "
                  f"Absolute times are unreliable; time differences "
                  f"within a sample remain valid.")

    df = df.sort_values(["ProbeNr", "Timestamp", "Nr"], kind="stable")

    if df["Timestamp"].notna().all():
        df["Minutes"] = (df.groupby("ProbeNr")["Timestamp"]
                           .transform(lambda s: (s - s.min()).dt.total_seconds() / 60))
    else:
        # Fallback index if timestamps cannot be parsed
        print("  WARNING: Timestamps incomplete; using measurement index as time axis.")
        df["Minutes"] = df.groupby("ProbeNr").cumcount()

    return df.reset_index(drop=True)


def calculate_composition(df: pd.DataFrame) -> pd.DataFrame:
    """Converts weighed masses [g] into mass fractions [wt.-%]."""
    df = df.copy()
    m_total = df[INITIAL_WEIGHTS].sum(axis=1)
    df["m_total"] = m_total
    safe_total = m_total.replace(0, np.nan)

    df["w_Al"]  = 100 * SL120["Al"] * df["m_SL120"] / safe_total
    df["w_IPA"] = 100 * (SL120["IPA"] * df["m_SL120"] + df["m_IPA"]) / safe_total
    df["w_PG"]  = 100 * (SL120["PG"] * df["m_SL120"] + df["m_PG"]) / safe_total
    df["w_MG"]  = 100 * df["m_MG"] / safe_total
    df["w_H2O"] = 100 * df["m_Wasser"] / safe_total

    for col in ["w_Al", "w_IPA", "w_PG", "w_MG", "w_H2O"]:
        df[col] = df[col].fillna(0.0)

    # Evaporation loss, if both initial and final masses were recorded
    has_loss_data = (df["m_vorher"] > 0) & (df["m_nachher"] > 0)
    df["Loss_pct"] = np.where(
        has_loss_data,
        100 * (df["m_vorher"] - df["m_nachher"]) / df["m_vorher"].replace(0, np.nan),
        np.nan
    )
    return df


def flag_unusable_records(df: pd.DataFrame) -> pd.DataFrame:
    """Marks records that must not be used for the plots.

    Three cases are caught:

    1. N below MIN_SAMPLES - the averaging buffer was empty or nearly
       empty when the record was written. All statistics are zero.
       Typical cause: the automatic write fired right after
       ResetStatistik.
    2. A measured value outside its plausible range - the sensor was
       not measuring liquid (air in the cell, station failure).
    3. NaN in one of the measured channels - Sound Vel returns NaN when
       no medium is present.

    Records are only flagged, never deleted. They remain in the Excel
    export together with the reason, so nothing disappears silently.
    """
    df = df.copy()
    reasons = pd.Series([""] * len(df), index=df.index, dtype=object)

    if "N" in df.columns:
        empty = df["N"].fillna(0) < MIN_SAMPLES
        reasons[empty] = reasons[empty] + f"N < {MIN_SAMPLES}; "

    for col, (lo, hi) in VALUE_RANGES.items():
        if col not in df.columns:
            continue
        values = pd.to_numeric(df[col], errors="coerce")
        bad = ~values.between(lo, hi) | values.isna()
        reasons[bad] = reasons[bad] + f"{col} outside {lo}..{hi}; "

    df["Exclusion_Reason"] = reasons.str.rstrip("; ")
    df["Record_Usable"] = df["Exclusion_Reason"] == ""
    return df


def detect_recipe_changes(df: pd.DataFrame) -> pd.DataFrame:
    """Identifies where initial weights have changed for each sample.

    During continuous recording with subsequent dosing, initial weights change
    in the middle of a sample run. These transition points should be highlighted
    in the plot.
    """
    df = df.copy()
    df["Recipe_Changed"] = False
    df["Change_Text"] = ""

    for probe, group in df.groupby("ProbeNr", sort=False):
        prev = group[INITIAL_WEIGHTS].shift()
        delta = group[INITIAL_WEIGHTS] - prev

        for idx in group.index[1:]:
            texts = []
            for col in INITIAL_WEIGHTS:
                d = delta.at[idx, col]
                if pd.notna(d) and abs(d) >= CHANGE_THRESHOLD:
                    texts.append(f"{d:+.1f} g {COLUMN_LABELS[col]}")
            if texts:
                df.at[idx, "Recipe_Changed"] = True
                df.at[idx, "Change_Text"] = ", ".join(texts)

    # Consecutive section index for each formulation/recipe step
    df["Recipe_Changed"] = df["Recipe_Changed"].fillna(False).astype(bool)
    df["Section"] = (df.groupby("ProbeNr")["Recipe_Changed"]
                       .cumsum().fillna(0).astype(int) + 1)
    return df


# =====================================================================
# 2  EXCEL EXPORT
# =====================================================================

def export_to_excel(df: pd.DataFrame) -> Path:
    """Writes processed measurement data and a summary per sample to .xlsx."""
    output_path = RESULTS_DIR / "measurement_data_processed.xlsx"

    columns = [
        "ProbeNr", "Section", "Nr", "Timestamp", "Minutes",
        "m_SL120", "m_Wasser", "m_IPA", "m_PG", "m_MG", "m_total",
        "w_Al", "w_IPA", "w_PG", "w_MG", "w_H2O",
        "Rho_M", "Rho_S", "Rho_Sp", "C_M", "C_S", "C_Sp",
        "T_M", "T_S", "Per_M", "Per_S", "RunT_M", "RunT_S",
        "N", "Stabil", "SensOK", "Gueltig",
        "Pumpe_pct", "m_vorher", "m_nachher", "Loss_pct",
        "Recipe_Changed", "Change_Text",
        "Record_Usable", "Exclusion_Reason", "Source_File",
    ]
    available_cols = [c for c in columns if c in df.columns]
    data = df[available_cols].copy()

    # Summary table grouped by sample ID and recipe section
    agg = {
        "Nr": "count", "Minutes": "max",
        "w_Al": "first", "w_IPA": "first", "w_PG": "first", "w_MG": "first",
        "Rho_M": ["mean", "std"], "C_M": ["mean", "std"],
        "T_M": ["mean", "min", "max"],
    }
    agg = {k: v for k, v in agg.items() if k in data.columns}
    summary = data.groupby(["ProbeNr", "Section"]).agg(agg)
    summary.columns = ["_".join(c).rstrip("_") for c in summary.columns]
    summary = summary.rename(columns={"Nr_count": "Record_Count",
                                      "Minutes_max": "Duration_min"})
    summary = summary.reset_index()

    with pd.ExcelWriter(output_path, engine="openpyxl",
                        datetime_format="YYYY-MM-DD HH:MM:SS") as xl:
        data.to_excel(xl, sheet_name="measurement_data", index=False)
        summary.to_excel(xl, sheet_name="Summary", index=False)
        _format_worksheet(xl.book["measurement_data"], len(available_cols))
        _format_worksheet(xl.book["Summary"], len(summary.columns))

    return output_path


def _format_worksheet(ws, num_cols: int) -> None:
    """Freezes header row, sets column widths, and standardizes font styling."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    data_font = Font(name="Arial")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    for i in range(1, num_cols + 1):
        col_letter = get_column_letter(i)
        max_len = max((len(str(cell.value)) for cell in ws[col_letter]
                       if cell.value is not None), default=8)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 9), 26)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =====================================================================
# 3  PLOTS
# =====================================================================

def _plot_channel(ax, x, mean_val, std_val, span_val, color, ylabel):
    """Plots a measurement channel with two dispersion/spread bands."""
    if span_val is not None:
        ax.fill_between(x, mean_val - span_val / 2, mean_val + span_val / 2,
                        color=color, alpha=0.12, lw=0, zorder=1)
    if std_val is not None:
        ax.fill_between(x, mean_val - Z_90 * std_val, mean_val + Z_90 * std_val,
                        color=color, alpha=0.28, lw=0, zorder=2)
    ax.plot(x, mean_val, "-", color=color, lw=1.6, zorder=3)
    ax.plot(x, mean_val, "o", color=color, ms=3.5, zorder=4)
    ax.set_ylabel(ylabel)
    ax.grid(alpha=0.35, lw=0.6, color=COLORS["grid"])


def _mark_changes(ax, change_x):
    for x in change_x:
        ax.axvline(x, color=COLORS["mark"], ls="--", lw=1.0, alpha=0.8, zorder=0)


def plot_sample(probe, sample_df: pd.DataFrame) -> Path:
    x = sample_df["Minutes"].to_numpy(float)
    changes = sample_df.loc[sample_df["Recipe_Changed"], "Minutes"].to_numpy(float)

    fig, axes = plt.subplots(4, 1, figsize=(12, 12), sharex=True,
                             gridspec_kw={"height_ratios": [1.0, 1.3, 1.3, 1.0]})

    # ---------- Panel 1: Composition as step function ----------
    ax = axes[0]
    has_plotted = False
    for name in ("Al", "IPA", "PG", "MG"):
        values = sample_df[f"w_{name}"].to_numpy(float)
        if np.nanmax(values) <= 0:
            continue
        ax.step(x, values, where="post", color=COLORS[name], lw=1.8, label=name)
        has_plotted = True
    if has_plotted:
        ax.legend(fontsize=8, ncol=4, loc="upper left", framealpha=0.9)
    else:
        ax.text(0.5, 0.5, "No active components – pure water",
                transform=ax.transAxes, ha="center", va="center", color="#777")
    ax.set_ylabel("Composition\n[wt.-%]")
    ax.grid(alpha=0.35, lw=0.6, color=COLORS["grid"])
    _mark_changes(ax, changes)

    # Annotations for recipe changes
    for _, row in sample_df[sample_df["Recipe_Changed"]].iterrows():
        ax.annotate(row["Change_Text"],
                    xy=(row["Minutes"], ax.get_ylim()[1]),
                    xytext=(4, -12), textcoords="offset points",
                    fontsize=7.5, rotation=90, va="top", ha="left",
                    color=COLORS["mark"])

    # ---------- Panels 2 and 3: Measured quantities ----------
    _plot_channel(axes[1], x, sample_df["Rho_M"].to_numpy(float),
                  sample_df["Rho_S"].to_numpy(float) if "Rho_S" in sample_df else None,
                  sample_df["Rho_Sp"].to_numpy(float) if "Rho_Sp" in sample_df else None,
                  COLORS["rho"], "Density\n[kg/m³]")
    _mark_changes(axes[1], changes)

    _plot_channel(axes[2], x, sample_df["C_M"].to_numpy(float),
                  sample_df["C_S"].to_numpy(float) if "C_S" in sample_df else None,
                  sample_df["C_Sp"].to_numpy(float) if "C_Sp" in sample_df else None,
                  COLORS["c"], "Sound Velocity\n[m/s]")
    _mark_changes(axes[2], changes)

    # ---------- Panel 4: Temperature ----------
    _plot_channel(axes[3], x, sample_df["T_M"].to_numpy(float),
                  sample_df["T_S"].to_numpy(float) if "T_S" in sample_df else None,
                  None, COLORS["T"], "Temperature\n[°C]")
    _mark_changes(axes[3], changes)
    axes[3].set_xlabel("Elapsed time since sample start [min]")

    # ---------- Legend for spread / uncertainty bands ----------
    handles = [
        Line2D([], [], color="#444", lw=1.6, marker="o", ms=4,
               label="Mean value over 100 measurements"),
        Line2D([], [], color="#444", lw=8, alpha=0.28,
               label="5% – 95% interval of single readings (±1.645·σ)"),
        Line2D([], [], color="#444", lw=8, alpha=0.12,
               label="Span (Min – Max) within window"),
        Line2D([], [], color=COLORS["mark"], ls="--", lw=1.0,
               label="Recipe / formulation change"),
    ]
    fig.legend(handles=handles, loc="lower center", ncol=4, fontsize=8,
               frameon=False, bbox_to_anchor=(0.5, -0.005))

    # ---------- Title ----------
    first_row = sample_df.iloc[0]
    header = (f"Sample {int(probe)}   ·   Start: "
              f"{first_row['w_Al']:.3f}% Al / {first_row['w_IPA']:.3f}% IPA / "
              f"{first_row['w_PG']:.3f}% PG / {first_row['w_MG']:.3f}% MG   ·   "
              f"Water balance {first_row['w_H2O']:.2f}%")
    if pd.notna(first_row.get("Timestamp", pd.NaT)):
        header += f"\nStart {first_row['Timestamp']:%Y-%m-%d %H:%M} (UTC)   ·   "
    else:
        header += "\n"
    header += (f"{len(sample_df)} records   ·   "
               f"{int(sample_df['Recipe_Changed'].sum())} recipe change(s)")
    fig.suptitle(header, fontsize=11, y=0.995)

    fig.tight_layout(rect=[0, 0.03, 1, 0.97])
    output_plot_path = RESULTS_DIR / f"Sample_{int(probe):03d}.png"
    fig.savefig(output_plot_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return output_plot_path


# =====================================================================
# MAIN PROGRAM
# =====================================================================

def main() -> None:
    for folder in (MEASUREMENT_DATA, RESULTS_DIR):
        folder.mkdir(exist_ok=True)

    print("=" * 68)
    print("DATA IMPORT & PREPROCESSING")
    print("=" * 68)
    df = load_measurement_data()
    df = add_time_columns(df)
    df = calculate_composition(df)
    df = flag_unusable_records(df)

    # Recipe changes are detected on the usable records only. Otherwise
    # a discarded record would sit between two real ones and the mass
    # delta would be split across it.
    usable = detect_recipe_changes(df[df["Record_Usable"]].copy())
    df = df.join(usable[["Recipe_Changed", "Change_Text", "Section"]])
    df["Recipe_Changed"] = df["Recipe_Changed"].fillna(False).astype(bool)
    df["Change_Text"] = df["Change_Text"].fillna("")
    df["Section"] = (df.groupby("ProbeNr")["Section"]
                       .ffill().bfill().fillna(1).astype(int))

    n_out = int((~df["Record_Usable"]).sum())
    print(f"  Total: {len(df)} records, "
          f"{df['ProbeNr'].nunique()} sample ID(s)")
    if n_out:
        print(f"  {n_out} record(s) excluded from the plots:")
        for _, row in df[~df["Record_Usable"]].iterrows():
            print(f"    Nr {int(row['Nr'])} (sample {int(row['ProbeNr'])}): "
                  f"{row['Exclusion_Reason']}")
        print("  They remain in the Excel export, column 'Record_Usable'.")

    print("\n" + "=" * 68)
    print("EXCEL EXPORT")
    print("=" * 68)
    excel_path = export_to_excel(df)
    print(f"  {excel_path.name}  (Sheets: measurement_data, Summary)")

    print("\n" + "=" * 68)
    print("GENERATING PLOTS")
    print("=" * 68)
    for probe, sample_df in df[df["Record_Usable"]].groupby("ProbeNr", sort=True):
        if sample_df.empty:
            print(f"  Sample {int(probe)}: no usable records, plot skipped")
            continue
        p = plot_sample(probe, sample_df.reset_index(drop=True))
        num_changes = int(sample_df["Recipe_Changed"].sum())
        change_info = f", {num_changes} recipe change(s)" if num_changes else ""
        dropped = int((~df.loc[df["ProbeNr"] == probe, "Record_Usable"]).sum())
        drop_info = f", {dropped} excluded" if dropped else ""
        print(f"  {p.name}  ({len(sample_df)} records"
              f"{change_info}{drop_info})")

    print("\nDone. All outputs are saved in 'results'.")


if __name__ == "__main__":
    main()